from flask import Flask, render_template, request, jsonify, send_from_directory
from flask_cors import CORS
import pdfplumber
import pandas as pd
import os
from datetime import datetime, timedelta
import re
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

app = Flask(__name__, static_folder='static')
CORS(app)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER']      = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def extract_flight_data_from_pdf(filepath, start_date_str, end_date_str):
    # Construir rango de fechas
    start_date = datetime.fromisoformat(start_date_str).date()
    end_date   = datetime.fromisoformat(end_date_str).date()
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    metadata = {
        'total_pages':     0,
        'extraction_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    headers = [
        "Línea Aérea", "Número Vuelo", "Procedencia", "Hora Llegada",
        "Número Vuelo Salida", "Destino", "Hora Salida",
        "Fecha de Operación"
    ]
    rows = []

    # Para primera y segunda pasada
    page_images   = {}
    header_labels = {}
    limits        = []
    day_headers_per_page = {}
    row_texts_per_page = {}

    with pdfplumber.open(filepath) as pdf:
        metadata['total_pages'] = len(pdf.pages)

        # 1ª pasada: extraer líneas de vuelo y posiciones de headers
        for pnum, page in enumerate(pdf.pages, start=1):
            start_idx = len(rows)
            text = page.extract_text(layout=True) or ""
            page_rows = []
            for line in text.split('\n'):
                if line.strip().startswith(('LAN','SKU','LXP')):
                    parts = re.split(r'\s{2,}', line.strip())
                    tokens = []
                    for p in parts:
                        tokens.extend(p.split(' '))
                    tokens = [t for t in tokens if t]
                    if len(tokens) > 8:
                        page_rows.append(tokens[:7])
                        page_rows.append(tokens[7:14])
                    elif len(tokens) >= 7:
                        page_rows.append(tokens[:7])
            rows.extend(page_rows)
            end_idx = len(rows)
            limits.append((pnum, start_idx, end_idx))
            page_images[pnum] = page.images
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            # Extraer headers de días de la semana (L, M, M, J, V, S, D)
            day_headers = [w for w in words if w['text'] in ['L','M','J','V','S','D'] or w['text'].startswith('M')]
            # Ordenar por posición x para mantener el orden visual
            day_headers = sorted(day_headers, key=lambda w: w['x0'])
            header_labels[pnum] = day_headers
            # Guardar los textos de los headers para el mapeo dinámico
            day_headers_per_page[pnum] = [w['text'] for w in day_headers]
            row_texts_per_page[pnum] = page_rows

        # 2ª pasada: mapear logos a offset de día y añadir fecha
        for pnum, start_idx, end_idx in limits:
            nrows = end_idx - start_idx
            if nrows <= 0:
                continue

            headers_ws = header_labels.get(pnum, [])
            day_cols   = [(w['x0'], w['x1']) for w in headers_ws]
            bottom_h   = max((w['bottom'] for w in headers_ws), default=0)

            # Mapeo dinámico: asociar cada columna de día a la fecha correcta del rango
            day_headers = day_headers_per_page.get(pnum, [])
            # Días de la semana en español (abreviados)
            dias_semana = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
            # Mapear headers a índices de día de la semana (0=Lunes, 1=Martes, ...)
            header_to_weekday = {}
            used_weekdays = set()
            for idx, h in enumerate(day_headers):
                # Resolver doble 'M' (Martes/Miércoles) por orden
                for widx, dw in enumerate(['L', 'M', 'M', 'J', 'V', 'S', 'D']):
                    if h == dw and widx not in used_weekdays:
                        header_to_weekday[idx] = widx
                        used_weekdays.add(widx)
                        break
            # Mapear columna a fecha del rango seleccionado
            col_to_date = {}
            for col_idx, weekday_idx in header_to_weekday.items():
                for d in date_range:
                    if d.weekday() == weekday_idx:
                        col_to_date[col_idx] = d
                        break

            page = pdf.pages[pnum-1]
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            # Mejorar: obtener los límites verticales de cada fila usando la posición de las palabras clave
            flight_rows = [w for w in words if w['text'] in ('LAN','SKU','LXP')]
            row_tops = sorted([w['top'] for w in flight_rows])
            row_bottoms = sorted([w['bottom'] for w in flight_rows])
            # Si hay menos filas detectadas por palabras clave que por parsing, igualar cantidad
            if len(row_tops) < nrows:
                # Interpolar límites
                min_top = min(row_tops) if row_tops else 0
                max_bottom = max(row_bottoms) if row_bottoms else (min_top + nrows*10)
                step = (max_bottom - min_top) / nrows if nrows else 1
                row_tops = [min_top + i*step for i in range(nrows)]
                row_bottoms = [min_top + (i+1)*step for i in range(nrows)]
            elif len(row_tops) > nrows:
                row_tops = row_tops[:nrows]
                row_bottoms = row_bottoms[:nrows]

            map_row = {i: [] for i in range(start_idx, end_idx)}
            for img in page_images[pnum]:
                cy = (img['top'] + img['bottom'])/2
                # Buscar la fila cuyo rango vertical contiene el centro del logo
                for ridx in range(nrows):
                    fila_top = row_tops[ridx]
                    fila_bottom = row_bottoms[ridx]
                    # Tolerancia: permitir un pequeño margen
                    if fila_top - 2 <= cy <= fila_bottom + 2:
                        map_row[ridx + start_idx].append(img)
                        break

            # Asignar fechas basadas en la posición de los logos en las columnas de día
            for i in range(start_idx, end_idx):
                logos = map_row.get(i, [])
                fechas = []
                for logo in logos:
                    cx = (logo['x0'] + logo['x1'])/2
                    centers = [ (x0+x1)/2 for x0,x1 in day_cols ]
                    if centers:
                        dc = min(range(len(centers)), key=lambda j: abs(cx-centers[j]))
                        fecha = col_to_date.get(dc)
                        if fecha:
                            dia_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][fecha.weekday()]
                            if f"{dia_semana}, {fecha.strftime('%Y-%m-%d')}" not in fechas:
                                fechas.append(f"{dia_semana}, {fecha.strftime('%Y-%m-%d')}")
                rows[i].append(fechas[0] if fechas else '')

    # Armar tabla de salida
    if rows:
        table = {
            'page':       1,
            'table_index':1,
            'headers':    headers,
            'data':       rows,
            'rows':       len(rows),
            'columns':    len(headers)
        }
        metadata['extracted_tables'] = 1
        metadata['extracted_rows']   = len(rows)
        return [table], metadata

    return [], metadata

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error':'No se envió archivo'}),400
        f = request.files['file']
        if f.filename == '':
            return jsonify({'error':'Nombre inválido'}),400
        if not allowed_file(f.filename):
            return jsonify({'error':'Extensión no permitida'}),400

        sd = request.form.get('start_date')
        ed = request.form.get('end_date')
        fn = secure_filename(f.filename)
        path = os.path.join(app.config['UPLOAD_FOLDER'], fn)
        f.save(path)

        tables, meta = extract_flight_data_from_pdf(path, sd, ed)
        if not tables:
            return jsonify({'warning':'No se encontraron datos','metadata':meta,'tables':[]}),200

        resp_meta = {
            'total_pages':     meta['total_pages'],
            'extracted_tables':meta['extracted_tables'],
            'extraction_date': meta['extraction_date'],
        }
        return jsonify({'success':True,'metadata':resp_meta,'tables':tables}),200

    except Exception as e:
        logger.error(f"upload_file error: {e}", exc_info=True)
        return jsonify({'error':f'Ocurrió un error: {e}'}),500

@app.route('/export/<file_format>', methods=['POST'])
def export_data(file_format):
    try:
        payload = request.get_json() or {}
        tables  = payload.get('tables', [])
        if not tables:
            return jsonify({'error':'No hay datos para exportar'}),400

        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_fn  = f'export_{ts}.{file_format}'
        out_path= os.path.join(app.config['UPLOAD_FOLDER'], out_fn)

        if file_format=='excel':
            # Ensure the file has .xlsx extension
            if not out_path.endswith('.xlsx'):
                out_path = os.path.splitext(out_path)[0] + '.xlsx'
                out_fn = os.path.splitext(out_fn)[0] + '.xlsx'
                
            # Create Excel writer with openpyxl engine
            with pd.ExcelWriter(out_path, engine='openpyxl', mode='w') as writer:
                for i, t in enumerate(tables):
                    df = pd.DataFrame(t['data'], columns=t['headers'])
                    name = f'Pag{t["page"]}_T{i+1}'[:31]  # Excel sheet name limit is 31 chars
                    df.to_excel(writer, sheet_name=name, index=False)
                
                # Save the workbook to apply styles
                writer._save()
                
            # Apply additional styling after saving
            style_excel_file(out_path, tables)

        elif file_format=='csv':
            dfs = [pd.DataFrame(t['data'], columns=t['headers']) for t in tables]
            pd.concat(dfs, ignore_index=True).to_csv(out_path, index=False, encoding='utf-8-sig')

        else:
            return jsonify({'error':'Formato no soportado'}),400

        return jsonify({'success':True,'filename':out_fn}),200

    except Exception as e:
        logger.error(f"export_data error: {e}", exc_info=True)
        return jsonify({'error':f'Error exportando: {e}'}),500

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

def style_excel_file(filepath, tables):
    wb = openpyxl.load_workbook(filepath)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'),   bottom=Side('thin'))
    for i, t in enumerate(tables):
        name = f'Pag{t["page"]}_T{i+1}'[:31]
        if name not in wb.sheetnames: continue
        ws = wb[name]
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = thin
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.border    = thin
                c.alignment = Alignment(wrap_text=True, horizontal='left')
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min((max_len+2)*1.2, 50)
        ws.freeze_panes = 'A2'
        tbl = Table(displayName=f"Table{i+1}", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
    wb.save(filepath)

if __name__=='__main__':
    app.run(debug=True, port=5000)
